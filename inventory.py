import openpyxl

# Load the workbook and select the sheet
excel_file   = openpyxl.load_workbook("inventory.xlsx")
product_list = excel_file["Sheet1"]

products_per_supplier    = {}      #Initialize the dictionary
total_value_per_supplier = {}
products_under_10_inv    = {}

for product_row in range(2, product_list.max_row + 1):              #Iterate through the rows (adjust the range if needed)
    supplier_name   = product_list.cell(product_row, 4).value         #Get supplier name from column 4
    inventory       = product_list.cell(product_row, 2).value
    price           = product_list.cell(product_row, 3).value
    product_num     = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    #Calculation number of products of supplier
    if supplier_name in products_per_supplier:                      
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1     #If product from the iterated row belong to a existing supplier(key), add into it's value
    else:
        print("Adding new supplier")                                                                #else add new supplier and initiate it's value from 1
        products_per_supplier[supplier_name] = 1


    #Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #Calculate products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    #Add value for total inventory price
    inventory_price.value = inventory * price

print(products_under_10_inv)
print(products_per_supplier)
print(total_value_per_supplier)                                                    #Print the dictionary

excel_file.save("inventory_with_total_value.xlsx")



