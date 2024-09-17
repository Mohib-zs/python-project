import openpyxl

# Load the workbook and select the sheet
excel_file = openpyxl.load_workbook("ExcelBook.xlsx")
employee_list = excel_file["Sheet1"]

# Initialize a dictionary to store the names and corresponding salaries
salary_per_employee = {}

# Iterate through the rows (adjust the range if needed)
for employee_row in range(4, employee_list.max_row - 5):
    # Get employee name from column 2 and salary from column 3
    employee_name = employee_list.cell(employee_row, 2).value
    salary = employee_list.cell(employee_row, 3).value

    # Add the employee's name and salary to the dictionary
    salary_per_employee[employee_name] = salary

# Print the dictionary containing names and salaries
print(salary_per_employee)
